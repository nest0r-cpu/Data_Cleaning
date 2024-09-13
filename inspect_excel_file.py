import openpyxl

def read_excel(file_path):
    # Load the Excel workbook
    try:
        workbook = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    # Print sheet names
    print("Sheet names:")
    for sheet in workbook.sheetnames:
        print(f"- {sheet}")

    # Print the first few rows of the first sheet
    sheet = workbook.active
    print("\nFirst few rows of the first sheet:")
    for row in sheet.iter_rows(values_only=True):
        print(row)
        if sheet.max_row < 5:
            break  # Limit to first 5 rows

if __name__ == "__main__":
    file_path = "Visualization_Movie_Data_Starter_Project.xlsx"  # Update this if necessary
    read_excel(file_path)
